import io

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

def parse_file(file_bytes, mime_type):
    if isinstance(file_bytes, bytes) and len(file_bytes) > MAX_FILE_SIZE:
        return f"Error: file too large ({len(file_bytes) // (1024*1024)}MB, max 10MB)"
    try:
        if "pdf" in mime_type:
            from PyPDF2 import PdfReader
            return "".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(file_bytes)).pages[:50])[:10000]
        elif "wordprocessing" in mime_type or "docx" in mime_type:
            from docx import Document
            return "\n".join(p.text for p in Document(io.BytesIO(file_bytes)).paragraphs)[:10000]
        elif "spreadsheet" in mime_type or "xlsx" in mime_type:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
            lines = []
            for s in wb.sheetnames[:3]:
                for row in wb[s].iter_rows(max_row=100, values_only=True):
                    lines.append(",".join(str(c or "") for c in row))
            return "\n".join(lines)[:10000]
        else: return file_bytes.decode("utf-8", errors="replace")[:10000]
    except Exception as e: return f"Error: failed to parse {mime_type}"
